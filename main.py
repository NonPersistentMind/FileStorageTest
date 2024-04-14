import os, asyncpg, typing as t, urllib
import urllib.parse
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse, FileResponse, Response, StreamingResponse
# from fastapi.middleware.cors import CORSMiddleware
from pathlib import Path
from openpyxl import Workbook as Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

import utils as u
from config import (
    DATA_FOLDER
)

DATABASE_URL: str = u.get_database_url()

app = FastAPI()
pool = None

async def startup_event() -> None:
    global pool
    pool = await asyncpg.create_pool(DATABASE_URL)
    await u.prepare_server()

async def shutdown_event() -> None:
    await pool.close()

app.add_event_handler("startup", startup_event)
app.add_event_handler("shutdown", shutdown_event)
        

@app.post("/files/{dir_name}")
async def upload_file_to_directory(dir_name: str, file: UploadFile = File(...)):
    try:
        file_path: Path = DATA_FOLDER / dir_name / file.filename
        file_path.parent.mkdir(parents=False, exist_ok=True)
        
        with open(file_path, "w+b") as f:
            bytes_written: int = f.write(await file.read())

        async with pool.acquire() as connection:
            file_id = await u.save_file_to_db(
                connection,
                dir_name,
                file.filename,
                bytes_written
            )
        return JSONResponse(content={"file_id": file_id}, status_code=201)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.post("/files")
async def upload_file(file: UploadFile = File(...)):
    try:
        file_path: Path = DATA_FOLDER / file.filename
        
        with open(file_path, "w+b") as f:
            bytes_written: int = f.write(await file.read())

        async with pool.acquire() as connection:
            file_id = await u.save_file_to_db(
                connection, 
                ".", # Save file to root directory
                file.filename,
                bytes_written
            )
        return JSONResponse(content={"file_id": file_id}, status_code=201)
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)


@app.get("/files/{file_id}")
async def download_file_by_id(file_id: int):
    async with pool.acquire() as connection:
        file_info = await u.get_file_info_by_id(connection, file_id)
        
    if not file_info:
        raise HTTPException(status_code=404, detail="File not found")

    file_name, dir_name, *_ = file_info
    file_path = DATA_FOLDER / dir_name / file_name

    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    response = FileResponse(path=file_path, headers={"Content-Disposition": f"attachment; filename={urllib.parse.quote(file_name.encode('utf-8'))}"})
    return response


@app.head("/files/{file_id}")
async def get_file_info_head(file_id: int):
    async with pool.acquire() as connection:
        file_info = await u.get_file_info_by_id(connection, file_id)
        
    if not file_info:
        raise HTTPException(status_code=404, detail="File not found")

    file_name, dir_name, file_size, updated_at  = file_info
    file_path = DATA_FOLDER / dir_name / file_name

    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    headers = {
        "Content-Length": str(file_size),
        "Content-Disposition": f"attachment; filename={urllib.parse.quote(file_name.encode('utf-8'))}",
        "X-File-Size": str(file_size),
        "X-Updated-At": updated_at.isoformat()
    }

    return Response(headers=headers)


@app.get("/top")
@app.get("/top/{dir_name}")
async def get_top_files(dir_name: str = None):
    async with pool.acquire() as connection:
        top_files_in_dir = await u.get_top_files(connection, dir_name)
    
    if not top_files_in_dir:
        raise HTTPException(status_code=404, detail="Directory not found")
    
    top_files_list = [
        {"file_name": row["file_name"], "file_size": row["file_size"], "updated_at": row["updated_at"].isoformat()}
        for row in top_files_in_dir
    ]

    return JSONResponse(content=top_files_list)


@app.get("/report/{file_id}")
async def generate_report(file_id: int):
    async with pool.acquire() as connection:
        file_info = await u.get_file_info_by_id(connection, file_id)
    if not file_info:
        raise HTTPException(status_code=404, detail="File not found")
    
    file_name, dir_name, *_ = file_info
    file_path: Path = DATA_FOLDER / dir_name / file_name

    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    file_data: str = file_path.read_text()
    try:
        rows = file_data.strip().split("\n")
        operations = rows[-1].split()[1:]  # Extract operation from the last row
        input_data = [row.split()[1:] for row in rows[:-1]]  # Extract input data from all but the last row

        wb = Workbook()
        ws = wb.active

        # Write all actual rows to the worksheet
        [ws.append(row.split()) for row in rows] 
        
        results = ["Результат:"]
        for col, operation in enumerate(operations):
            col_letter = get_column_letter(col+2)
            # Combine the operation with the input data into a formula
            results.append(f"={operation.join([f'{col_letter+str(row)}' for row in range(1, len(input_data)+1)])}") 
        ws.append(results)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={urllib.parse.quote((file_name+'. Report.xlsx').encode('utf-8'))}"
            }
        )
    except Exception as e:
        return JSONResponse(content={"error": "File format is not appropriate"}, status_code=400)

